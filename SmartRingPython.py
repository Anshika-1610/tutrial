import serial
import openpyxl
from twilio.rest import Client
from serial import Serial
from playsound import playsound
ser=serial.Serial('COM3',9600)
data=ser.readline().decode('ascii')
iD=str(data[0])
wb=openpyxl.load_workbook("C:\Book1.xlsx")
sheet=wb.get_sheet_by_name("Sheet1")
for i in range(1,10): #We have taken a dataset of only 10 entries the iteration can be extended upto any number
    if(sheet.cell(row=i,column=1).value==1):
        name=str(sheet.cell(row=i,column=2).value)
        add=str(sheet.cell(row=i,column=3).value)
        p=str(sheet.cell(row=i,column=4).value)
        img=str(sheet.cell(row=i,column=5).value)
pno="+91"+p
print(name)
print(add)
print(pno)
print(img)
account_sid = 'The account sid is confidential'
auth_token = 'The auth Token is confidential'
client = Client(account_sid, auth_token)

message = client.messages \
                .create(
                     body='''Emergency!!!
*Here will be the link to our web page when it get hosted on web*''',
                     from_='+19382226885',
                     to=pno
                 )
print(message.sid)
playsound("C:\\salamisound-5863157-yelp-signal-for-rescue-u-s.mp3")
var=1
while var==1:
    data=ser.readline().decode('ascii')
    a="https://www.google.com/maps/place/"
    b=data[1:22]
    c=a+b
    f = open("C:\\Users\\Aryan Varshney\\Desktop\\file_1.html","w")
    f.write('''<html lang="en">
    <head>
        <meta http-equiv="refresh" content="10">
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
        <link rel="stylesheet" href="C:\\Users\Aryan Varshney\\Desktop\\style.css">
        <title>Emergency Ring</title>
    </head>
    <body><center>

        <img src="download.png"  width=15% height=20% ">
        <table>
            <tr>
              <th>Image</th>
              <th>Name</th>
              <th>I.D.</th>
              <th>Address</th>
              <th>Phone No.</th>
            </tr>
            <tr>
              <td><img src = "'''+img+'''" style="width:100px; height:100px"></td>
              <td>'''+name+'''</td>
              <td>'''+iD+'''</td>
              <td>'''+add+'''</td>
              <td>'''+pno+'''</td>
            </tr>
        </table>
    <br>
    <iframe width="1000" height="500" src="'''+c+'''">
    </iframe>
    </center>
    </body>
    </html>''')
    f.close()
    